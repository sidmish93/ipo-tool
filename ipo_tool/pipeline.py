"""
IPO shareholding Excel pipeline.

Given a date range, this:
  1. Scrapes SEBI "Final Offer Documents filed with ROC" for companies in the range.
  2. Looks up each on BSE (scrip code, tickers, ISIN, market cap).
  3. Keeps companies with Mcap Full > 3000 cr.
  4. Finds the latest shareholding quarter + statement page links.
  5. Extracts detailed promoter and public shareholding.
  6. Writes a two-sheet Excel workbook.

Everything that touches SEBI / BSE / NSE runs inside a real Chromium browser
(via Playwright) because SEBI blocks plain HTTP POSTs and the public-shareholder
"bold vs non-bold" distinction only exists in the rendered DOM.
"""

import re
import time
import datetime
import urllib.parse

from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36")

SEBI_LIST = ("https://www.sebi.gov.in/sebiweb/home/HomeAction.do"
             "?doListing=yes&sid=3&ssid=15&smid=12")
MCAP_THRESHOLD = 3000.0


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _norm(name):
    """Normalise a company name for fuzzy matching."""
    if not name:
        return ""
    s = name.lower()
    s = re.sub(r"[^a-z0-9]", "", s)
    for suf in ("privatelimited", "limited", "ltd", "pvt"):
        if s.endswith(suf):
            s = s[: -len(suf)]
    return s


def _to_float(txt):
    if txt is None:
        return None
    s = str(txt).replace(",", "").strip()
    if s in ("", "-", "NA", "N.A."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


_MONTHS = {m[:3].lower(): m for m in
           ["January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"]}


def _display_quarter(qname):
    """Normalise a BSE quarter label to 'Month YYYY'."""
    if not qname:
        return ""
    qname = qname.strip()
    m = re.match(r"^\d{1,2}[-/ ]([A-Za-z]{3,})[-/ ](\d{2,4})$", qname)
    if m:
        mon = _MONTHS.get(m.group(1)[:3].lower(), m.group(1).title())
        yr = m.group(2)
        if len(yr) == 2:
            yr = "20" + yr
        return f"{mon} {yr}"
    return qname


def _clean_company(title):
    """'Turtlemint Fintech Solutions Limited - Prospectus' -> company name.

    Some SME rows contain two lines (Prospectus + Abridged Prospectus); keep the
    first line only, then strip the trailing '- Prospectus' marker.
    """
    first_line = title.splitlines()[0] if title else ""
    t = re.sub(r"\s*-\s*(abridged\s+)?prospectus.*$", "", first_line,
               flags=re.IGNORECASE).strip()
    return t


# --------------------------------------------------------------------------- #
# Browser-evaluated JS snippets
# --------------------------------------------------------------------------- #
_SEBI_ROWS_JS = r"""
() => {
  const rows = [];
  document.querySelectorAll('table tr').forEach(tr => {
    const td = tr.querySelectorAll('td');
    if (td.length >= 2) {
      const d = (td[0].innerText || '').trim();
      const t = (td[1].innerText || '').trim();
      if (t && /\d{4}/.test(d)) rows.push([d, t]);
    }
  });
  return rows;
}
"""

# --------------------------------------------------------------------------- #
# Pipeline
# --------------------------------------------------------------------------- #
class Pipeline:
    def __init__(self, progress_cb=None):
        self.cb = progress_cb or (lambda *a, **k: None)

    def log(self, msg, current=None, total=None, stage=None):
        self.cb(msg, current=current, total=total, stage=stage)

    # ----- SEBI ----------------------------------------------------------- #
    def _sebi_companies(self, page, dfrom, dto):
        """Return [(name, date)] of prospectuses filed within [dfrom, dto]."""
        page.goto(SEBI_LIST, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(1500)
        f_str = dfrom.strftime("%d-%m-%Y")
        t_str = dto.strftime("%d-%m-%Y")
        page.evaluate(
            "(v)=>{document.getElementById('fromDate').value=v.f;"
            "document.getElementById('toDate').value=v.t;}",
            {"f": f_str, "t": t_str},
        )

        def trigger(js):
            try:
                with page.expect_response(
                    lambda r: "HomeAction.do" in r.url
                    and r.request.method == "POST",
                    timeout=25000,
                ):
                    page.evaluate(js)
            except Exception:
                page.evaluate(js)
            page.wait_for_timeout(1200)

        all_rows = []
        prev_first = None
        for page_idx in range(0, 25):
            if page_idx == 0:
                trigger("()=>searchFormNewsList('s','-1')")
            else:
                trigger(f"()=>searchFormNewsList('n','{page_idx}')")
            rows = page.evaluate(_SEBI_ROWS_JS)
            if not rows:
                break
            first = tuple(rows[0])
            if page_idx > 0 and first == prev_first:
                break  # pager clamped -> no more pages
            all_rows.extend(rows)
            prev_first = first
            self.log(f"SEBI page {page_idx + 1}: {len(rows)} rows",
                     stage="sebi")
            has_next = page.evaluate(
                r"""() => {
                  let n=false;
                  document.querySelectorAll('a').forEach(a=>{
                    const t=(a.innerText||'').trim().toLowerCase();
                    const h=a.getAttribute('href')||'';
                    if((t==='next'||t==='last')&&/searchFormNewsList/.test(h)) n=true;
                  });
                  return n;
                }"""
            )
            if not has_next:
                break

        # parse + filter by date + dedupe
        out = []
        seen = set()
        for d, t in all_rows:
            try:
                dt = datetime.datetime.strptime(d.strip(), "%b %d, %Y").date()
            except ValueError:
                continue
            if not (dfrom <= dt <= dto):
                continue
            name = _clean_company(t)
            key = _norm(name)
            if not key or key in seen:
                continue
            seen.add(key)
            out.append((name, dt))
        return out

    # ----- BSE ------------------------------------------------------------ #
    _FETCH_JS = """async (u) => {
        try {
          const r = await fetch(u, {headers:{'Accept':'application/json'}});
          const t = await r.text();
          try { return {ok:true, json:JSON.parse(t)}; }
          catch(e){ return {ok:false, text:t.slice(0,200)}; }
        } catch(e){ return {ok:false, text:''+e}; }
      }"""

    def _bse_fetch(self, bpage, url, retries=3):
        for attempt in range(retries):
            try:
                return bpage.evaluate(self._FETCH_JS, url)
            except Exception:
                try:
                    bpage.wait_for_timeout(700)
                except Exception:
                    pass
        return {"ok": False, "text": "evaluate failed"}

    @staticmethod
    def _seg_rank(type_str):
        """Rank BSE search segments: Equity T+1 is the one to use.

        A company can appear under 'Equity T+1', 'Derivatives' and 'Equity T+0'.
        The T+0 row has a *different* scrip code (e.g. 143529 / DELHIVERY#) whose
        shareholding pattern is blank, so it must be avoided in favour of T+1.
        """
        t = (type_str or "").lower()
        if "equity t+1" in t:
            return 0
        if "equity t+0" in t:
            return 3
        if "deriv" in t:
            return 4
        if "equity" in t:
            return 1
        return 2

    @staticmethod
    def _name_rank(scripname, target):
        n = _norm(scripname)
        if n and n == target:
            return 0
        if n and (n.startswith(target) or target.startswith(n)
                  or target in n or n in target):
            return 1
        return 2

    def _bse_lookup(self, bpage, name):
        """Return dict with scripcode/ticker/isin/bse_name or None.

        A company can show up in several segments. Searching the *full* name
        ("X Limited") often returns ONLY the Equity T+0 phantom (scrip like
        143529 / 'X#') whose shareholding pattern is blank, while the real
        Equity T+1 listing ('X Ltd', 543529) only appears for a shorter query.
        So we collect hits across all query variants and pick the best name
        match in the Equity T+1 segment (see _seg_rank / _name_rank).
        """
        queries = [name, re.sub(r"\s+(limited|ltd)\.?$", "", name,
                                flags=re.IGNORECASE)]
        words = name.split()
        if len(words) >= 3:
            queries.append(" ".join(words[:3]))
        if len(words) >= 2:
            queries.append(" ".join(words[:2]))

        target = _norm(name)
        seen, cands = set(), []

        def rank(r):
            return (self._name_rank(r.get("scripName", ""), target),
                    self._seg_rank(r.get("Type", "")))

        for q in queries:
            url = ("https://api.bseindia.com/BseIndiaAPI/api/"
                   "GetQuoteAllSearchDatabeta/w?searchString="
                   + urllib.parse.quote(q))
            res = self._bse_fetch(bpage, url)
            if res.get("ok") and isinstance(res["json"], list):
                for r in res["json"]:
                    key = (str(r.get("strSricpCode", "")), r.get("Type", ""))
                    if key not in seen:
                        seen.add(key)
                        cands.append(r)
            if cands:
                best = min(cands, key=rank)
                nr, sr = rank(best)
                # stop only once we have a solid name match in a real equity
                # segment (T+1 or plain equity) — never on a T+0/derivative row
                if nr <= 1 and sr <= 1:
                    return self._pack_lookup(best)

        if cands:
            return self._pack_lookup(min(cands, key=rank))
        return None

    @staticmethod
    def _pack_lookup(r):
        return {
            "scripcode": str(r.get("strSricpCode", "")).strip(),
            "ticker": (r.get("shortName") or "").strip(),
            "isin": (r.get("Isin") or "").strip(),
            "bse_name": (r.get("scripName") or "").strip(),
        }

    def _bse_mcap(self, bpage, scripcode):
        url = ("https://api.bseindia.com/BseIndiaAPI/api/StockTrading/w"
               "?flag=&quotetype=EQ&scripcode=" + scripcode)
        res = self._bse_fetch(bpage, url)
        if res.get("ok"):
            return _to_float(res["json"].get("MktCapFull"))
        return None

    def _bse_quarter(self, bpage, scripcode):
        url = ("https://api.bseindia.com/BseIndiaAPI/api/"
               "CorporatesSHPSecuritybeta/w?scripcode=" + scripcode + "&qtrid=")
        res = self._bse_fetch(bpage, url)
        if not res.get("ok"):
            return None
        tbl = res["json"].get("Table") or []
        if not tbl:
            return None
        row = tbl[0]
        qid = row.get("Qtr_Id")
        qname = row.get("Fld_qtrname") or ""
        if qid is None:
            return None
        return {"qid": qid, "qname": qname}

    # ----- NSE ------------------------------------------------------------ #
    def _nse_ticker(self, npage, name, bse_name, fallback):
        query = re.sub(r"\s+(limited|ltd)\.?$", "", name, flags=re.IGNORECASE)
        url = ("https://www.nseindia.com/api/NextApi/globalSearch/equity?symbol="
               + urllib.parse.quote(query))
        targets = {_norm(name), _norm(bse_name)}
        for attempt in range(3):
            try:
                res = npage.evaluate(self._FETCH_JS, url)
            except Exception:
                # NSE occasionally reloads for its bot cookie; re-settle and retry
                try:
                    npage.goto("https://www.nseindia.com/",
                               wait_until="domcontentloaded")
                    npage.wait_for_timeout(1500)
                except Exception:
                    pass
                continue
            if res.get("ok"):
                data = (res["json"] or {}).get("data") or []
                eq = [d for d in data
                      if (d.get("series") or "").upper() == "EQ"] or data
                for d in eq:
                    if _norm(d.get("companyName", "")) in targets:
                        return (d.get("symbol") or "").strip()
                if eq:
                    return (eq[0].get("symbol") or "").strip()
                return fallback
            npage.wait_for_timeout(800)
        return fallback

    # ----- statement links + detailed shareholding ------------------------ #
    @staticmethod
    def _stmt_urls(scripcode, qid, qname):
        """Human-viewable statement pages (used as clickable links in Sheet 1)."""
        q = f"{float(qid):.2f}"
        enc = urllib.parse.quote(qname)
        base = "https://www.bseindia.com/corporates/"
        return (
            f"{base}ShpPromoterNGroup?scripcd={scripcode}&qtrid={q}&QtrName={enc}",
            f"{base}shpPublicShareholder?scripcd={scripcode}&qtrid={q}&QtrName={enc}",
        )

    @staticmethod
    def _largest_table(j):
        big, blen = None, -1
        for k, v in (j or {}).items():
            if isinstance(v, list) and len(v) > blen:
                big, blen = k, len(v)
        return (j.get(big) if big else []) or []

    def _promoter_rows(self, bpage, scripcode, qid):
        """[name, 'Promoter'/'Promoter Group', pct] where pct(% A+B+C2) > 0."""
        url = ("https://api.bseindia.com/BseIndiaAPI/api/"
               f"Corp_shpPromoterNGroup_ng/w?SCRIPCODE={scripcode}"
               f"&QtrCode={float(qid):.2f}")
        res = self._bse_fetch(bpage, url)
        out = []
        if res.get("ok"):
            for x in self._largest_table(res["json"]):
                ty = (x.get("FLd_ShareholderType") or "").strip()
                if ty not in ("Promoter", "Promoter Group"):
                    continue
                pct = _to_float(x.get("Fld_TotalPercentageOf_A_B_C2"))
                if pct is None or pct <= 0:
                    continue
                nm = re.sub(r"\s+", " ", (x.get("Fld_ShareHolderName") or "").strip())
                out.append([nm, ty, pct])
        return out

    def _public_rows(self, bpage, scripcode, qid):
        """[name, bold-heading, pct] for every non-bold public holder, pct > 0."""
        url = ("https://api.bseindia.com/BseIndiaAPI/api/"
               f"Corp_shpSec_SHPPubShold_ng/w?SCRIPCODE={scripcode}"
               f"&QtrCode={float(qid):.2f}")
        res = self._bse_fetch(bpage, url)
        out = []
        if res.get("ok"):
            for x in self._largest_table(res["json"]):
                nm = re.sub(r"\s+", " ", (x.get("Fld_ShareHolderName") or "").strip())
                if not nm:
                    continue
                pct = _to_float(x.get("Fld_TotalPercentageOf_A_B_C2"))
                if pct is None or pct <= 0:
                    continue
                heading = (x.get("Fld_Level") or x.get("Fld_SubCategory") or "").strip()
                heading = re.sub(r"/+\s*$", "", heading).strip()
                out.append([nm, heading, pct])
        return out

    # ----- orchestration -------------------------------------------------- #
    def run(self, dfrom, dto, out_path):
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    # flags that keep Chromium stable / lean on small cloud VMs
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-gpu",
                    "--disable-extensions",
                ],
            )
            ctx = browser.new_context(user_agent=UA,
                                      viewport={"width": 1440, "height": 900})
            ctx.set_default_timeout(60000)
            try:
                sebi_page = ctx.new_page()
                self.log("Fetching SEBI final offer documents...", stage="sebi")
                companies = self._sebi_companies(sebi_page, dfrom, dto)
                self.log(f"Found {len(companies)} companies filed with ROC in range.",
                         stage="sebi")
                sebi_page.close()

                bpage = ctx.new_page()
                bpage.goto("https://www.bseindia.com/", wait_until="domcontentloaded")
                bpage.wait_for_timeout(1500)

                npage = ctx.new_page()
                try:
                    npage.goto("https://www.nseindia.com/",
                               wait_until="domcontentloaded")
                    npage.wait_for_timeout(1500)
                except Exception:
                    pass

                qualified = []
                total = len(companies)
                for i, (name, dt) in enumerate(companies, 1):
                    self.log(f"[{i}/{total}] {name}", current=i, total=total,
                             stage="bse")
                    info = self._bse_lookup(bpage, name)
                    if not info or not info["scripcode"]:
                        self.log(f"    not found on BSE, skipping", stage="bse")
                        continue
                    mcap = self._bse_mcap(bpage, info["scripcode"])
                    if mcap is None or mcap <= MCAP_THRESHOLD:
                        self.log(f"    Mcap {mcap} <= {MCAP_THRESHOLD:.0f}cr, skipping",
                                 stage="bse")
                        continue
                    qtr = self._bse_quarter(bpage, info["scripcode"])
                    if not qtr:
                        self.log("    no shareholding quarter, skipping", stage="bse")
                        continue
                    nse = self._nse_ticker(npage, name, info["bse_name"],
                                           info["ticker"])
                    prom_url, pub_url = self._stmt_urls(
                        info["scripcode"], qtr["qid"], qtr["qname"])
                    promoters = self._promoter_rows(
                        bpage, info["scripcode"], qtr["qid"])
                    public = self._public_rows(
                        bpage, info["scripcode"], qtr["qid"])
                    qualified.append({
                        "name": name,
                        "bse_ticker": info["ticker"],
                        "scripcode": info["scripcode"],
                        "nse_ticker": nse,
                        "isin": info["isin"],
                        "mcap": mcap,
                        "quarter": _display_quarter(qtr["qname"]),
                        "prom_url": prom_url,
                        "pub_url": pub_url,
                        "promoters": promoters,
                        "public": public,
                    })
                    self.log(f"    QUALIFIED  Mcap {mcap:,.2f}cr  {info['ticker']}  "
                             f"(promoters={len(promoters)}, public={len(public)})",
                             stage="bse")
                    bpage.wait_for_timeout(150)

                bpage.close()
                npage.close()
            finally:
                ctx.close()
                browser.close()

        self.log("Writing Excel workbook...", stage="excel")
        _write_workbook(qualified, out_path)
        self.log(f"Done. {len(qualified)} companies written.", stage="done")
        return qualified

    # ----- single-company lookup (interactive search) -------------------- #
    def lookup_one(self, name):
        """Look up one company by name and return its full shareholding detail.

        Unlike run(), there is no market-cap filter here — it returns whatever
        the company is, so it works for interactive search of any listed name.
        """
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--no-sandbox", "--disable-dev-shm-usage",
                    "--disable-gpu", "--disable-extensions",
                ],
            )
            ctx = browser.new_context(user_agent=UA,
                                      viewport={"width": 1440, "height": 900})
            ctx.set_default_timeout(60000)
            try:
                bpage = ctx.new_page()
                bpage.goto("https://www.bseindia.com/",
                           wait_until="domcontentloaded")
                bpage.wait_for_timeout(1200)

                info = self._bse_lookup(bpage, name)
                if not info or not info["scripcode"]:
                    return {"found": False, "query": name}

                mcap = self._bse_mcap(bpage, info["scripcode"])
                qtr = self._bse_quarter(bpage, info["scripcode"])

                npage = ctx.new_page()
                try:
                    npage.goto("https://www.nseindia.com/",
                               wait_until="domcontentloaded")
                    npage.wait_for_timeout(1200)
                    nse = self._nse_ticker(npage, info["bse_name"],
                                           info["bse_name"], info["ticker"])
                except Exception:
                    nse = info["ticker"]
                finally:
                    npage.close()

                promoters, public, prom_url, pub_url, quarter = [], [], "", "", ""
                if qtr:
                    quarter = _display_quarter(qtr["qname"])
                    prom_url, pub_url = self._stmt_urls(
                        info["scripcode"], qtr["qid"], qtr["qname"])
                    promoters = self._promoter_rows(
                        bpage, info["scripcode"], qtr["qid"])
                    public = self._public_rows(
                        bpage, info["scripcode"], qtr["qid"])

                return {
                    "found": True,
                    "query": name,
                    "name": info["bse_name"],
                    "bse_ticker": info["ticker"],
                    "scripcode": info["scripcode"],
                    "nse_ticker": nse,
                    "isin": info["isin"],
                    "mcap": mcap,
                    "quarter": quarter,
                    "prom_url": prom_url,
                    "pub_url": pub_url,
                    "promoters": promoters,
                    "public": public,
                }
            finally:
                ctx.close()
                browser.close()


# --------------------------------------------------------------------------- #
# Excel writer
# --------------------------------------------------------------------------- #
def _write_workbook(companies, out_path):
    wb = Workbook()

    # ---- Sheet 1 ---- #
    ws = wb.active
    ws.title = "IPO Companies >3000cr"
    headers = ["S.No", "Company Name", "BSE Ticker", "BSE Scrip Code",
               "NSE Ticker", "ISIN", "Mcap Full (Cr.)",
               "Latest Shareholding Quarter",
               "Promoter & Promoter Group Shareholding (link)",
               "Public Shareholder Shareholding (link)"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    link_font = Font(color="0563C1", underline="single")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for i, comp in enumerate(companies, 1):
        r = i + 1
        vals = [i, comp["name"], comp["bse_ticker"], comp["scripcode"],
                comp["nse_ticker"], comp["isin"], comp["mcap"], comp["quarter"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = left if c == 2 else center
            if c == 7 and isinstance(v, (int, float)):
                cell.number_format = "#,##0.00"
        pc = ws.cell(row=r, column=9, value="Promoter & Promoter Group Statement")
        pc.hyperlink = comp["prom_url"]
        pc.font = link_font
        pc.alignment = left
        pc.border = border
        uc = ws.cell(row=r, column=10, value="Public Shareholder Statement")
        uc.hyperlink = comp["pub_url"]
        uc.font = link_font
        uc.alignment = left
        uc.border = border

    widths = [6, 42, 14, 12, 14, 16, 16, 16, 40, 36]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    if companies:
        ws.auto_filter.ref = f"A1:J{len(companies) + 1}"

    # ---- Sheet 2 ---- #
    ws2 = wb.create_sheet("Detailed Shareholding")
    h2 = ["Company Name", "Statement", "Category / Name of Shareholder",
          "Promoter Type / Public Heading",
          "Shareholding % as per SCRR,1957 (% of A+B+C2)"]
    prom_fill = PatternFill("solid", fgColor="E2EFDA")
    pub_fill = PatternFill("solid", fgColor="FCE4D6")
    comp_font = Font(bold=True, size=11, color="1F4E78")
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    def wrow(r, fill, vals, aligns):
        for c, (v, al) in enumerate(zip(vals, aligns), 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.alignment = al
            cell.fill = fill
            cell.border = border
            if c == 5 and isinstance(v, (int, float)):
                cell.number_format = "0.00"

    r = 2
    for comp in companies:
        start = r
        prom = comp.get("promoters") or []
        pub = comp.get("public") or []
        if prom:
            for nm, ty, pct in prom:
                wrow(r, prom_fill,
                     [comp["name"], "Promoter & Promoter Group", nm, ty, pct],
                     [left, left, left, center, center])
                r += 1
        else:
            wrow(r, prom_fill,
                 [comp["name"], "Promoter & Promoter Group",
                  "No promoter / promoter group with shareholding > 0%",
                  "-", "-"], [left, left, left, center, center])
            r += 1
        if pub:
            for nm, hd, pct in pub:
                wrow(r, pub_fill,
                     [comp["name"], "Public Shareholder", nm, hd, pct],
                     [left, left, left, left, center])
                r += 1
        else:
            wrow(r, pub_fill,
                 [comp["name"], "Public Shareholder",
                  "No named (non-bold) public shareholder with shareholding > 0%",
                  "-", "-"], [left, left, left, center, center])
            r += 1
        ws2.cell(row=start, column=1).font = comp_font

    for c, w in enumerate([40, 26, 52, 40, 22], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.freeze_panes = "A2"
    if r > 2:
        ws2.auto_filter.ref = f"A1:E{r - 1}"

    wb.save(out_path)


def run_pipeline(from_str, to_str, out_path, progress_cb=None):
    """from_str / to_str are 'YYYY-MM-DD' (HTML date input format)."""
    dfrom = datetime.datetime.strptime(from_str, "%Y-%m-%d").date()
    dto = datetime.datetime.strptime(to_str, "%Y-%m-%d").date()
    return Pipeline(progress_cb).run(dfrom, dto, out_path)


def lookup_company(name):
    """Interactive single-company lookup; returns a JSON-serialisable dict."""
    return Pipeline().lookup_one(name)
